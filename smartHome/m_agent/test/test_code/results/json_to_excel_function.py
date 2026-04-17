# -*- coding: utf-8 -*-
"""
将实验结果目录批量转为 Excel，适合在 PyCharm 中作为函数调用。

目录结构示例：
project_root/
├── json_to_excel_function.py
├── ourAgent/
│   ├── gemini-2.5-flash/
│   │   ├── 1_smart_home_test_results.json
│   │   ├── 2_smart_home_test_results.json
│   │   └── 3_smart_home_test_results.json
│   ├── gpt-5-mini/
│   │   ├── 1_smart_home_test_results.json
│   │   ├── 2_smart_home_test_results.json
│   │   └── 3_smart_home_test_results.json
│   └── gpt-5-nano/
│       ├── 1_smart_home_test_results.json
│       ├── 2_smart_home_test_results.json
│       └── 3_smart_home_test_results.json
├── sage/
└── sasha/

功能说明：
1. 每个 agent 生成一个 Excel 文件：
   - ourAgent_experiment_results.xlsx
   - sage_experiment_results.xlsx
   - sasha_experiment_results.xlsx

2. 每个 Excel 文件包含 3 个工作簿：
   - gpt5mini
   - gpt5nano
   - gemini

3. 每个工作簿列结构：
   - 第 1 列：测试用例（格式：用例编号.测试用例字符串）
   - 第 2 列：第 1 次结果
   - 第 3 列：第 2 次结果
   - 第 4 列：第 3 次结果

4. 每个用例结果使用 T / F 表示。

在 PyCharm 中调用示例：
    from json_to_excel_function import export_all_agents_to_excels

    export_all_agents_to_excels(base_dir=".")
"""

import json
import re
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


MODEL_FOLDER_TO_SHEET = OrderedDict([
    ("gpt-5-mini", "gpt5mini"),
    ("gpt-5-nano", "gpt5nano"),
    ("gemini-2.5-flash", "gemini"),
])

AGENT_FOLDERS = ["ourAgent", "sage", "sasha"]


def _to_tf_text(value):
    """把各种可能的结果值规范成 T / F。"""
    if isinstance(value, bool):
        return "T" if value else "F"

    if isinstance(value, (int, float)):
        return "T" if bool(value) else "F"

    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"true", "t", "1", "yes", "y", "通过", "正确"}:
            return "T"
        if v in {"false", "f", "0", "no", "n", "未通过", "错误"}:
            return "F"

    return "F"


def _extract_run_index(file_path):
    """
    从文件名中提取实验轮次。
    例如：
        1_smart_home_test_results.json -> 1
        2_xxx.json -> 2
    """
    match = re.match(r"(\d+)_", file_path.name)
    if match:
        return int(match.group(1))
    return 10**9


def _load_single_run(json_path):
    """
    读取一次实验结果，返回：
        OrderedDict({
            "1.Network status": "T",
            "2.Turn off light": "F",
            ...
        })
    """
    with Path(json_path).open("r", encoding="utf-8") as f:
        data = json.load(f)

    details = data.get("详细结果", [])
    if not isinstance(details, list):
        raise ValueError(f"文件格式错误，'详细结果' 必须是列表：{json_path}")

    run_results = OrderedDict()
    for item in details:
        case_id = item.get("用例编号", "")
        case_text = item.get("测试用例字符串", "")
        case_name = f"{case_id}.{case_text}"
        run_results[case_name] = _to_tf_text(item.get("是否正确", False))

    return run_results


def _merge_runs_to_rows(run_dicts):
    """
    将 3 次实验结果合并成表格行数据。

    参数：
        run_dicts: [run1_dict, run2_dict, run3_dict]

    返回：
        rows = [
            ["1.xxx", "T", "F", "T"],
            ["2.xxx", "F", "F", "T"],
            ...
        ]
    """
    all_cases = OrderedDict()
    for run_dict in run_dicts:
        for case_name in run_dict.keys():
            all_cases.setdefault(case_name, None)

    rows = []
    for case_name in all_cases.keys():
        row = [case_name]
        for run_dict in run_dicts:
            row.append(run_dict.get(case_name, ""))
        rows.append(row)

    return rows


def _apply_sheet_style(ws, data_row_count):
    """统一设置工作簿样式。"""
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 数据区样式
    for row in ws.iter_rows(min_row=2, max_row=max(2, data_row_count + 1), min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    if data_row_count > 0:
        ws.auto_filter.ref = f"A1:D{data_row_count + 1}"


def _write_model_sheet(ws, rows):
    """写入单个模型工作簿，并在底部统计每次实验的通过数和正确率。"""
    ws["A1"] = "测试用例"
    ws["B1"] = "第1次"
    ws["C1"] = "第2次"
    ws["D1"] = "第3次"

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 底部统计
    data_count = len(rows)
    pass_count_row = data_count + 3
    accuracy_row = data_count + 4

    pass_count_1 = sum(1 for row in rows if len(row) > 1 and row[1] == "T")
    pass_count_2 = sum(1 for row in rows if len(row) > 2 and row[2] == "T")
    pass_count_3 = sum(1 for row in rows if len(row) > 3 and row[3] == "T")

    accuracy_1 = pass_count_1 / data_count if data_count > 0 else 0
    accuracy_2 = pass_count_2 / data_count if data_count > 0 else 0
    accuracy_3 = pass_count_3 / data_count if data_count > 0 else 0

    ws.cell(row=pass_count_row, column=1, value="通过数")
    ws.cell(row=pass_count_row, column=2, value=pass_count_1)
    ws.cell(row=pass_count_row, column=3, value=pass_count_2)
    ws.cell(row=pass_count_row, column=4, value=pass_count_3)

    ws.cell(row=accuracy_row, column=1, value="正确率")
    ws.cell(row=accuracy_row, column=2, value=accuracy_1)
    ws.cell(row=accuracy_row, column=3, value=accuracy_2)
    ws.cell(row=accuracy_row, column=4, value=accuracy_3)

    ws.cell(row=accuracy_row, column=2).number_format = "0.00%"
    ws.cell(row=accuracy_row, column=3).number_format = "0.00%"
    ws.cell(row=accuracy_row, column=4).number_format = "0.00%"

    # 平均正确率（三次正确率加和 / 3，保留 1 位小数，直接截断）
    average_accuracy = (accuracy_1 + accuracy_2 + accuracy_3) / 3
    average_accuracy_percent_truncated = int(average_accuracy * 1000) / 10

    ws.cell(row=accuracy_row, column=5, value="平均正确率")
    ws.cell(row=accuracy_row, column=6, value=f"{average_accuracy_percent_truncated:.1f}%")

    _apply_sheet_style(ws, len(rows))

    # 统计区样式
    summary_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for row in [pass_count_row, accuracy_row]:
        max_col = 6 if row == accuracy_row else 4
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = summary_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def export_single_agent_to_excel(agent_dir, output_excel_path):
    """
    为单个 agent 生成一个 Excel 文件。

    参数：
        agent_dir: agent 目录，例如 "./ourAgent"
        output_excel_path: 输出 Excel 路径，例如 "./ourAgent_experiment_results.xlsx"

    返回：
        输出 Excel 的绝对路径（字符串）
    """
    agent_dir = Path(agent_dir)
    output_excel_path = Path(output_excel_path)

    if not agent_dir.exists():
        raise FileNotFoundError(f"agent 目录不存在：{agent_dir}")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for model_folder, sheet_name in MODEL_FOLDER_TO_SHEET.items():
        model_dir = agent_dir / model_folder
        if not model_dir.exists():
            raise FileNotFoundError(f"模型目录不存在：{model_dir}")

        json_files = sorted(
            model_dir.glob("*_smart_home_test_results.json"),
            key=_extract_run_index
        )

        if len(json_files) == 0:
            run_dicts = [OrderedDict(), OrderedDict(), OrderedDict()]
        else:
            run_dicts = [_load_single_run(json_file) for json_file in json_files[:3]]
            while len(run_dicts) < 3:
                run_dicts.append(OrderedDict())

        rows = _merge_runs_to_rows(run_dicts)

        ws = wb.create_sheet(title=sheet_name)
        _write_model_sheet(ws, rows)

    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel_path)
    return str(output_excel_path.resolve())


def export_all_agents_to_excels(base_dir="."):
    """
    批量为所有 agent 生成 Excel 文件。

    参数：
        base_dir: 根目录，默认当前目录

    返回：
        dict，形如：
        {
            "ourAgent": "/abs/path/ourAgent_experiment_results.xlsx",
            "sage": "/abs/path/sage_experiment_results.xlsx",
            "sasha": "/abs/path/sasha_experiment_results.xlsx",
        }
    """
    base_dir = Path(base_dir)
    output_paths = {}

    for agent_name in AGENT_FOLDERS:
        agent_dir = base_dir / agent_name
        output_excel_path = base_dir / "experiment_results" / f"{agent_name}_experiment_results.xlsx"
        output_paths[agent_name] = export_single_agent_to_excel(
            agent_dir=agent_dir,
            output_excel_path=output_excel_path
        )

    return output_paths


if __name__ == "__main__":
    # 在当前目录下批量生成 3 个 agent 的 Excel 文件
    result = export_all_agents_to_excels(base_dir=".")
    for agent_name, excel_path in result.items():
        print(f"{agent_name}: {excel_path}")
