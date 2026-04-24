import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def load_result_file(file_path: Path):
    with file_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    result_map = {}
    for item in data.get("详细结果", []):
        case_id = item["用例编号"]
        result_map[case_id] = {
            "测试用例字符串": item.get("测试用例字符串", ""),
            "是否正确": bool(item.get("是否正确", False))
        }
    return result_map


def main():
    base_dir = Path(".")
    ouragent_dir = base_dir / "ourAgent"
    output_dir = base_dir / "failure_results"
    output_dir.mkdir(parents=True, exist_ok=True)

    output_file = output_dir / "ourAgent_three_run_failures.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    model_order = ["gpt-5-mini", "gpt-5-nano", "gemini-2.5-flash"]

    # 样式
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    summary_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    bold_font = Font(bold=True)

    header_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for model_name in model_order:
        model_dir = ouragent_dir / model_name
        if not model_dir.exists():
            continue

        run_data = []
        for i in range(1, 4):
            result_file = model_dir / f"{i}_smart_home_test_results.json"
            run_data.append(load_result_file(result_file))

        all_case_ids = sorted(run_data[0].keys())

        ws = wb.create_sheet(title=model_name[:31])
        ws.append(["测试用例（编号+测试用例字符串）", "是否连错三次"])

        # 表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = header_alignment

        fail_count = 0

        for case_id in all_case_ids:
            test_str = run_data[0][case_id]["测试用例字符串"]

            failed_all_three = all(
                run.get(case_id, {}).get("是否正确", False) is False
                for run in run_data
            )

            if failed_all_three:
                fail_count += 1

            case_text = f"{case_id} - {test_str}"
            mark = "T" if failed_all_three else ""

            ws.append([case_text, mark])

            current_row = ws.max_row
            ws.cell(current_row, 1).alignment = wrap_alignment
            ws.cell(current_row, 2).alignment = center_alignment

        # 最后一行统计
        ws.append(["连错三次总数", fail_count])
        summary_row = ws.max_row

        for cell in ws[summary_row]:
            cell.fill = summary_fill
            cell.font = bold_font
            cell.alignment = center_alignment

        # 列宽
        ws.column_dimensions["A"].width = 85
        ws.column_dimensions["B"].width = 18

        # 冻结表头
        ws.freeze_panes = "A2"

    wb.save(output_file)
    print(f"已生成: {output_file}")


if __name__ == "__main__":
    main()
